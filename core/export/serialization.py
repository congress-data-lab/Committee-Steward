"""Serialization helpers for canonical release exports."""

from __future__ import annotations

import csv
import hashlib
import json
import re
import zipfile
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook.properties import CalcProperties

from .schema import (
    DATASET_SPECS,
    METADATA_FIELDS_EXCLUDED_FROM_SEMANTIC_HASH,
    RELEASE_METADATA_SHEET,
    WORKBOOK_SHEET_ORDER,
    ColumnSpec,
    DatasetSpec,
)


DETERMINISTIC_ZIP_TIMESTAMP = (2000, 1, 1, 0, 0, 0)
DETERMINISTIC_WORKBOOK_TIMESTAMP = "2000-01-01T00:00:00Z"
WORKBOOK_CREATOR = "Committee Steward"


def _canonicalize_workbook_member(filename: str, payload: bytes) -> bytes:
    """Remove save-time volatility from an OOXML workbook member."""
    if filename != "docProps/core.xml":
        return payload

    canonical, replacement_count = re.subn(
        rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
        rb"\g<1>" + DETERMINISTIC_WORKBOOK_TIMESTAMP.encode("ascii") + rb"\g<2>",
        payload,
    )
    if replacement_count != 1:
        raise ValueError("Expected exactly one modified timestamp in workbook core properties")
    return canonical


def render_cell_value(value: Any, data_type: str) -> Any:
    """Normalize typed values before writing or hashing."""
    if value is None:
        return None
    if data_type == "date":
        if isinstance(value, datetime):
            return value.date()
        return value
    if data_type == "boolean":
        return bool(value)
    if data_type == "integer":
        return int(value)
    if data_type == "float":
        if isinstance(value, Decimal):
            value = float(value)
        return round(float(value), 3)
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return str(value)


def serialize_value_for_csv(value: Any, data_type: str) -> str:
    normalized = render_cell_value(value, data_type)
    if normalized is None:
        return ""
    if data_type == "date":
        return normalized.isoformat()
    if data_type == "boolean":
        return "true" if normalized else "false"
    if data_type == "float":
        return f"{normalized:.3f}"
    return str(normalized)


def serialize_rows_for_csv(spec: DatasetSpec, rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    serialized: list[dict[str, str]] = []
    for row in rows:
        serialized.append(
            {
                column.name: serialize_value_for_csv(row.get(column.name), column.data_type)
                for column in spec.columns
            }
        )
    return serialized


def _write_csv(path: Path, spec: DatasetSpec, rows: list[dict[str, Any]]) -> str:
    serialized = serialize_rows_for_csv(spec, rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.stem + ".tmp" + path.suffix)
    with temporary.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[column.name for column in spec.columns],
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(serialized)
    temporary.replace(path)
    return sha256_file(path)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 64), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _semantic_cell_payload(
    value: Any,
    column: ColumnSpec,
    *,
    metadata_field: str | None = None,
) -> dict[str, Any]:
    normalized = render_cell_value(value, column.data_type)
    if metadata_field in METADATA_FIELDS_EXCLUDED_FROM_SEMANTIC_HASH:
        normalized = ""
    if normalized is None:
        rendered = None
    elif column.data_type == "date":
        rendered = normalized.isoformat()
    else:
        rendered = normalized
    return {"type": column.data_type, "value": rendered}


def compute_semantic_workbook_hash(sheet_rows: dict[str, list[dict[str, Any]]]) -> str:
    payload: list[dict[str, Any]] = []
    for sheet_name in WORKBOOK_SHEET_ORDER:
        matching_spec = next(spec for spec in DATASET_SPECS.values() if spec.sheet_name == sheet_name)
        rows = sheet_rows[sheet_name]
        semantic_rows = []
        for row in rows:
            metadata_field = row.get("field") if sheet_name == RELEASE_METADATA_SHEET else None
            semantic_rows.append(
                [
                    _semantic_cell_payload(row.get(column.name), column, metadata_field=metadata_field)
                    for column in matching_spec.columns
                ]
            )
        payload.append(
            {
                "sheet_name": sheet_name,
                "headers": [column.name for column in matching_spec.columns],
                "rows": semantic_rows,
            }
        )
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _set_string_cell(cell: Cell, value: str) -> None:
    cell.value = value
    if value.startswith("="):
        cell.data_type = "s"


def write_workbook(path: Path, sheet_rows: dict[str, list[dict[str, Any]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = WORKBOOK_CREATOR
    workbook.properties.lastModifiedBy = WORKBOOK_CREATOR
    workbook.properties.created = datetime(2000, 1, 1, 0, 0, 0)
    workbook.properties.modified = datetime(2000, 1, 1, 0, 0, 0)
    workbook.calculation = CalcProperties(calcMode="manual", fullCalcOnLoad=False, forceFullCalc=False)

    for sheet_name in WORKBOOK_SHEET_ORDER:
        spec = next(item for item in DATASET_SPECS.values() if item.sheet_name == sheet_name)
        worksheet = workbook.create_sheet(title=sheet_name)
        headers = [column.name for column in spec.columns]
        worksheet.append(headers)
        worksheet.freeze_panes = "A2"
        for row in sheet_rows[sheet_name]:
            worksheet.append([render_cell_value(row.get(column.name), column.data_type) for column in spec.columns])
        worksheet.auto_filter.ref = worksheet.dimensions
        for row_index, row in enumerate(sheet_rows[sheet_name], start=2):
            for column_index, column in enumerate(spec.columns, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                value = row.get(column.name)
                if value is None:
                    continue
                if column.data_type == "string":
                    _set_string_cell(cell, str(value))
                elif column.data_type == "date":
                    cell.number_format = "yyyy-mm-dd"

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.stem + ".tmp" + path.suffix)
    workbook.save(temporary)
    normalized = path.with_name(path.stem + ".normalized" + path.suffix)
    with zipfile.ZipFile(temporary, "r") as source, zipfile.ZipFile(
        normalized,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as destination:
        for source_info in sorted(source.infolist(), key=lambda item: item.filename):
            info = zipfile.ZipInfo(
                source_info.filename,
                date_time=DETERMINISTIC_ZIP_TIMESTAMP,
            )
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = (
                (0o40755 if source_info.is_dir() else 0o100644) << 16
            )
            payload = _canonicalize_workbook_member(
                source_info.filename,
                source.read(source_info.filename),
            )
            destination.writestr(info, payload)
    normalized.replace(temporary)
    temporary.replace(path)


def write_release_artifacts(
    output_dir: Path,
    file_rows: dict[str, list[dict[str, Any]]],
    *,
    congress_from: int,
    congress_to: int,
    release_metadata: dict[str, Any],
) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    range_suffix = f"{congress_from}_{congress_to}"

    written_files: dict[str, dict[str, Any]] = {}
    sheet_rows = {
        DATASET_SPECS["assignments"].sheet_name: file_rows["assignments"],
        DATASET_SPECS["rankings"].sheet_name: file_rows["rankings"],
        DATASET_SPECS["events"].sheet_name: file_rows["events"],
        DATASET_SPECS["members"].sheet_name: file_rows["members"],
        DATASET_SPECS["committees"].sheet_name: file_rows["committees"],
        DATASET_SPECS["sources"].sheet_name: file_rows["sources"],
        DATASET_SPECS["validation"].sheet_name: file_rows["validation"],
        DATASET_SPECS["data_dictionary"].sheet_name: file_rows["data_dictionary"],
        DATASET_SPECS["release_metadata"].sheet_name: file_rows["release_metadata"],
    }

    for key in (
        "assignments",
        "rankings",
        "events",
        "members",
        "committees",
        "sources",
        "validation",
        "directory_mismatches",
    ):
        spec = DATASET_SPECS[key]
        filename = f"{spec.filename_stem}_{range_suffix}.csv"
        path = output_dir / filename
        sha256 = _write_csv(path, spec, file_rows[key])
        written_files[filename] = {"sha256": sha256, "rows": len(file_rows[key])}

    workbook_filename = f"committee_membership_{range_suffix}.xlsx"
    workbook_path = output_dir / workbook_filename
    write_workbook(workbook_path, sheet_rows)
    written_files[workbook_filename] = {
        "sha256": sha256_file(workbook_path),
        "rows": sum(len(rows) for rows in sheet_rows.values()),
    }

    metadata_path = output_dir / "release_metadata.json"
    metadata_path.write_text(
        json.dumps(release_metadata, sort_keys=True, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    written_files[metadata_path.name] = {"sha256": sha256_file(metadata_path), "rows": None}

    sha256sums_path = output_dir / "SHA256SUMS"
    lines = [
        f"{info['sha256']}  {filename}"
        for filename, info in sorted(written_files.items())
    ]
    sha256sums_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    written_files[sha256sums_path.name] = {"sha256": sha256_file(sha256sums_path), "rows": None}
    return written_files


def extract_sheet_rows_from_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    extracted: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in WORKBOOK_SHEET_ORDER:
        spec = next(item for item in DATASET_SPECS.values() if item.sheet_name == sheet_name)
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        headers = list(rows[0])
        expected_headers = [column.name for column in spec.columns]
        if headers != expected_headers:
            raise AssertionError(f"Unexpected headers for {sheet_name}: {headers!r}")
        normalized_rows: list[dict[str, Any]] = []
        for values in rows[1:]:
            row: dict[str, Any] = {}
            for column, value in zip(spec.columns, values):
                if value == "":
                    value = None
                if isinstance(value, datetime) and column.data_type == "date":
                    value = value.date()
                row[column.name] = value
            normalized_rows.append(row)
        extracted[sheet_name] = normalized_rows
    return extracted
