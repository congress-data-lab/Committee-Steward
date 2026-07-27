from __future__ import annotations

import zipfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from core.export.release import (
    SCHEMA_VERSION,
    _stable_now,
    build_assignment_rows,
    build_directory_mismatch_rows,
    build_data_dictionary_rows,
    build_event_rows,
    build_ranking_rows,
    build_release_metadata_rows,
    build_source_rows,
    is_standing_committee,
    make_release_event_id,
    make_release_source_document_id,
    make_release_source_id,
)
from core.export.schema import DATASET_SPECS, WORKBOOK_SHEET_ORDER
from core.export.serialization import (
    compute_semantic_workbook_hash,
    extract_sheet_rows_from_workbook,
    serialize_rows_for_csv,
    write_release_artifacts,
    write_workbook,
)
from validate.directory_overlap import DirectoryMismatch


def _sample_sheet_rows() -> dict[str, list[dict]]:
    release_source_id = make_release_source_id(
        source_type="resolution",
        source_name="House Resolution",
        version_tag="2023-01-03",
    )
    release_source_document_id = make_release_source_document_id(
        release_source_id=release_source_id,
        external_id="HRES-1",
        doc_date=date(2023, 1, 3),
        url="https://example.test/HRES-1",
        content_hash="abc123",
    )
    appointment_event_id = make_release_event_id(
        congress_no=118,
        chamber="H",
        bioguide_id="A000001",
        committee_code="HSAG",
        action="APPOINTED",
        decision_date=date(2023, 1, 3),
        effective_date=date(2023, 1, 3),
        release_source_document_id=release_source_document_id,
        source_locator="p. H1",
    )
    removal_event_id = make_release_event_id(
        congress_no=118,
        chamber="H",
        bioguide_id="A000001",
        committee_code="HSAG",
        action="REMOVED",
        decision_date=date(2023, 8, 1),
        effective_date=date(2023, 8, 1),
        release_source_document_id=release_source_document_id,
        source_locator="p. H100",
    )
    assignments = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "congress_no": 118,
            "chamber": "H",
            "bioguide_id": "A000001",
            "committee_code": "HSAG",
            "committee_name": "Committee on Agriculture",
            "start_date": date(2023, 1, 3),
            "last_active_date": date(2023, 7, 31),
            "termination_effective_date": date(2023, 8, 1),
            "ended_early": True,
            "start_release_event_id": appointment_event_id,
            "end_release_event_id": removal_event_id,
            "internal_start_event_id": "db-1",
            "internal_end_event_id": "db-2",
        }
    ]
    rankings = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "congress_no": 118,
            "chamber": "H",
            "bioguide_id": "A000001",
            "committee_code": "HSAG",
            "committee_name": "Committee on Agriculture",
            "caucus_party_code": 100,
            "rank_in_party": 1,
            "unresolved_slots_before": 0,
            "rank_start_date": date(2023, 1, 3),
            "rank_last_active_date": date(2025, 1, 2),
            "rank_end_boundary": date(2025, 1, 3),
            "rank_basis": "resolution_order",
            "rank_observation_id": "rank-1",
            "release_source_document_id": release_source_document_id,
            "source_locator": "HRES-1#appointment[0]",
            "raw_member_name": "Ms. Example",
            "rank_after_raw_name": None,
            "observation_kind": "FULL_ROSTER",
        }
    ]
    events = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "release_event_id": appointment_event_id,
            "internal_event_id": "db-1",
            "congress_no": 118,
            "chamber": "H",
            "bioguide_id": "A000001",
            "committee_code": "HSAG",
            "committee_name": "Committee on Agriculture",
            "action": "APPOINTED",
            "decision_date": date(2023, 1, 3),
            "effective_date": date(2023, 1, 3),
            "release_source_document_id": release_source_document_id,
            "internal_source_document_id": 11,
            "source_locator": "p. H1",
            "text_span": "Resolved, That Alice Example be appointed.",
            "extraction_mode": "resolution_structured",
            "note_types": None,
            "interpretation_basis": None,
        },
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "release_event_id": removal_event_id,
            "internal_event_id": "db-2",
            "congress_no": 118,
            "chamber": "H",
            "bioguide_id": "A000001",
            "committee_code": "HSAG",
            "committee_name": "Committee on Agriculture",
            "action": "REMOVED",
            "decision_date": date(2023, 8, 1),
            "effective_date": date(2023, 8, 1),
            "release_source_document_id": release_source_document_id,
            "internal_source_document_id": 12,
            "source_locator": "p. H100",
            "text_span": "Without objection, Alice Example was excused.",
            "extraction_mode": "record_pattern",
            "note_types": "explicit_resignation",
            "interpretation_basis": "House floor request",
        },
    ]
    members = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "bioguide_id": "A000001",
            "congress_no": 118,
            "chamber": "H",
            "service_start": date(2023, 1, 3),
            "service_last_active_date": date(2025, 1, 2),
            "first_name": "Alice",
            "last_name": "Example",
            "official_full_name": "Alice Example",
            "nickname": None,
            "state": "NY",
            "district": 1,
            "party_code": 100,
            "caucus_party_code": 100,
            "exit_reason": None,
        }
    ]
    committees = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "committee_code": "HSAG",
            "chamber": "H",
            "committee_name": "Committee on Agriculture",
            "valid_start": date(2011, 1, 5),
            "valid_last_active_date": None,
            "is_joint": False,
        }
    ]
    sources = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "release_source_id": release_source_id,
            "internal_source_id": 5,
            "release_source_document_id": release_source_document_id,
            "internal_source_document_id": 11,
            "source_type": "resolution",
            "source_name": "House Resolution",
            "version_tag": "2023-01-03",
            "external_id": "HRES-1",
            "doc_date": date(2023, 1, 3),
            "url": "https://example.test/HRES-1",
            "content_hash": "abc123",
            "retrieved_at_utc": "2026-07-14T00:00:00Z",
            "created_at_utc": "2026-07-14T00:00:00Z",
        }
    ]
    validation = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "validation_policy_version": "1.0.0",
            "validation_type": "directory_overlap",
            "congress_no": 118,
            "chamber": "H",
            "snapshot_date": date(2023, 2, 1),
            "committee_scope": "standing",
            "event_type": None,
            "gate_status": "PASS",
            "reference_count": 1,
            "observed_count": 1,
            "overlap_count": 1,
            "reference_only_count": 0,
            "observed_only_count": 0,
            "directory_member_entries": 1,
            "resolved_directory_assignments": 1,
            "unresolved_member_entries": 0,
            "unmapped_committee_entries": 0,
            "unmapped_committees": 0,
            "member_resolution_pct": 100.0,
            "reference_coverage_pct": None,
            "directory_coverage_pct": 100.0,
            "observed_overlap_pct": 100.0,
            "date_comparable_count": None,
            "exact_date_match_count": None,
            "exact_date_match_pct": None,
            "within_one_day_match_count": None,
            "within_one_day_match_pct": None,
        }
    ]
    metadata = build_release_metadata_rows(
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "generated_at_utc": "2026-07-14T00:00:00Z",
            "workbook_semantic_sha256": "",
            "scope": "standing",
        }
    )
    return {
        "Assignments": assignments,
        "Rankings": rankings,
        "Events": events,
        "Members": members,
        "Committees": committees,
        "Sources": sources,
        "Validation": validation,
        "Data Dictionary": build_data_dictionary_rows(congress_from=118, congress_to=118),
        "Release Metadata": metadata,
    }


def test_build_ranking_rows_uses_half_open_date_semantics() -> None:
    rows = build_ranking_rows(
        [
            {
                "congress_no": 118,
                "chamber": "H",
                "bioguide_id": "A000001",
                "committee_code": "HSAG",
                "committee_name": "Committee on Agriculture",
                "caucus_party_code": 100,
                "rank_in_party": 3,
                "unresolved_slots_before": 1,
                "rank_start_date": date(2023, 1, 3),
                "rank_end_boundary": date(2023, 8, 1),
                "rank_basis": "relative_instruction",
                "rank_observation_id": "rank-1",
                "release_source_document_id": "source-document-1",
                "source_locator": "HRES-1#appointment[0]",
                "raw_member_name": "Ms. Example",
                "rank_after_raw_name": "Mr. Anchor",
                "observation_kind": "RELATIVE_ORDER",
            }
        ],
        release_version="v0.2.0",
    )

    assert rows[0]["rank_last_active_date"] == date(2023, 7, 31)
    assert rows[0]["rank_in_party"] == 3
    assert rows[0]["unresolved_slots_before"] == 1


def test_release_source_ids_are_invariant_when_internal_ids_change() -> None:
    rows = build_source_rows(
        [
            {
                "internal_source_document_id": 11,
                "internal_source_id": 5,
                "release_source_id": make_release_source_id(
                    source_type="resolution",
                    source_name="Senate Resolution",
                    version_tag="2023-01-04",
                ),
                "release_source_document_id": make_release_source_document_id(
                    release_source_id=make_release_source_id(
                        source_type="resolution",
                        source_name="Senate Resolution",
                        version_tag="2023-01-04",
                    ),
                    external_id="SRES-1",
                    doc_date=date(2023, 1, 4),
                    url="https://example.test/SRES-1",
                    content_hash="same-hash",
                ),
                "source_type": "resolution",
                "source_name": "Senate Resolution",
                "version_tag": "2023-01-04",
                "external_id": "SRES-1",
                "doc_date": date(2023, 1, 4),
                "url": "https://example.test/SRES-1",
                "content_hash": "same-hash",
                "retrieved_at_utc": datetime(2026, 7, 14, 12, 0, 0),
                "created_at_utc": datetime(2026, 7, 14, 12, 0, 0),
            },
            {
                "internal_source_document_id": 99,
                "internal_source_id": 77,
                "release_source_id": make_release_source_id(
                    source_type="resolution",
                    source_name="Senate Resolution",
                    version_tag="2023-01-04",
                ),
                "release_source_document_id": make_release_source_document_id(
                    release_source_id=make_release_source_id(
                        source_type="resolution",
                        source_name="Senate Resolution",
                        version_tag="2023-01-04",
                    ),
                    external_id="SRES-1",
                    doc_date=date(2023, 1, 4),
                    url="https://example.test/SRES-1",
                    content_hash="same-hash",
                ),
                "source_type": "resolution",
                "source_name": "Senate Resolution",
                "version_tag": "2023-01-04",
                "external_id": "SRES-1",
                "doc_date": date(2023, 1, 4),
                "url": "https://example.test/SRES-1",
                "content_hash": "same-hash",
                "retrieved_at_utc": datetime(2026, 7, 14, 12, 0, 0),
                "created_at_utc": datetime(2026, 7, 14, 12, 0, 0),
            },
        ],
        release_version="v0.1.0",
    )

    assert rows[0]["release_source_id"] == rows[1]["release_source_id"]
    assert rows[0]["release_source_document_id"] == rows[1]["release_source_document_id"]
    assert rows[0]["internal_source_id"] != rows[1]["internal_source_id"]
    assert rows[0]["internal_source_document_id"] != rows[1]["internal_source_document_id"]


def test_release_event_ids_are_invariant_to_internal_ids_but_distinguish_same_day_evidence() -> None:
    release_source_id = make_release_source_id(
        source_type="resolution",
        source_name="Senate Resolution",
        version_tag="2023-01-04",
    )
    release_source_document_id = make_release_source_document_id(
        release_source_id=release_source_id,
        external_id="SRES-1",
        doc_date=date(2023, 1, 4),
        url="https://example.test/SRES-1",
        content_hash="same-hash",
    )
    raw_base = {
        "congress_no": 118,
        "chamber": "S",
        "bioguide_id": "B000002",
        "committee_code": "SSAF",
        "committee_name": "Committee on Agriculture, Nutrition, and Forestry",
        "action": "APPOINTED",
        "decision_date": date(2023, 1, 4),
        "effective_date": date(2023, 1, 4),
        "release_source_document_id": release_source_document_id,
        "note_types": None,
        "interpretation_basis": None,
    }
    invariant_rows, _ = build_event_rows(
        [
            {
                **raw_base,
                "internal_event_id": "db-a",
                "internal_source_document_id": 11,
                "source_locator": "p. S1",
                "text_span": "Ordered, That Bob Example be appointed.",
                "extraction_mode": "resolution_structured",
            },
            {
                **raw_base,
                "internal_event_id": "db-b",
                "internal_source_document_id": 99,
                "source_locator": "p. S1",
                "text_span": "Ordered, That Bob Example be appointed.",
                "extraction_mode": "resolution_structured",
            },
        ],
        release_version="v0.1.0",
    )
    assert invariant_rows[0]["release_event_id"] == invariant_rows[1]["release_event_id"]

    distinct_rows, _ = build_event_rows(
        [
            {
                **raw_base,
                "internal_event_id": "db-c",
                "internal_source_document_id": 11,
                "source_locator": "p. S1",
                "text_span": "Ordered, That Bob Example be appointed.",
                "extraction_mode": "resolution_structured",
            },
            {
                **raw_base,
                "internal_event_id": "db-d",
                "internal_source_document_id": 11,
                "source_locator": "p. S2",
                "text_span": "Ordered, That Bob Example be appointed.",
                "extraction_mode": "journal_pattern",
            },
        ],
        release_version="v0.1.0",
    )
    assert len(distinct_rows) == 2
    assert distinct_rows[0]["release_event_id"] != distinct_rows[1]["release_event_id"]


def test_assignments_use_release_ids_and_distinct_date_semantics() -> None:
    release_source_id = make_release_source_id(
        source_type="resolution",
        source_name="House Resolution",
        version_tag="2023-01-03",
    )
    release_source_document_id = make_release_source_document_id(
        release_source_id=release_source_id,
        external_id="HRES-1",
        doc_date=date(2023, 1, 3),
        url="https://example.test/HRES-1",
        content_hash="abc123",
    )
    event_rows, event_id_map = build_event_rows(
        [
            {
                "internal_event_id": "db-1",
                "congress_no": 118,
                "chamber": "H",
                "bioguide_id": "A000001",
                "committee_code": "HSAG",
                "committee_name": "Committee on Agriculture",
                "action": "APPOINTED",
                "decision_date": date(2023, 1, 3),
                "effective_date": date(2023, 1, 3),
                "release_source_document_id": release_source_document_id,
                "internal_source_document_id": 11,
                "source_locator": "p. H1",
                "text_span": "Resolved, That Alice Example be appointed.",
                "extraction_mode": "resolution_structured",
                "note_types": None,
                "interpretation_basis": None,
            },
            {
                "internal_event_id": "db-2",
                "congress_no": 118,
                "chamber": "H",
                "bioguide_id": "A000001",
                "committee_code": "HSAG",
                "committee_name": "Committee on Agriculture",
                "action": "REMOVED",
                "decision_date": date(2023, 8, 1),
                "effective_date": date(2023, 8, 1),
                "release_source_document_id": release_source_document_id,
                "internal_source_document_id": 12,
                "source_locator": "p. H100",
                "text_span": "Without objection, Alice Example was excused.",
                "extraction_mode": "record_pattern",
                "note_types": None,
                "interpretation_basis": None,
            },
            {
                "internal_event_id": "db-3",
                "congress_no": 118,
                "chamber": "S",
                "bioguide_id": "B000002",
                "committee_code": "SSAF",
                "committee_name": "Committee on Agriculture, Nutrition, and Forestry",
                "action": "APPOINTED",
                "decision_date": date(2023, 1, 3),
                "effective_date": date(2023, 1, 3),
                "release_source_document_id": release_source_document_id,
                "internal_source_document_id": 13,
                "source_locator": "p. S1",
                "text_span": "Ordered, That Bob Example be appointed.",
                "extraction_mode": "resolution_structured",
                "note_types": None,
                "interpretation_basis": None,
            },
        ],
        release_version="v0.1.0",
    )
    assert event_rows[0]["release_event_id"] == event_id_map["db-1"]

    rows = build_assignment_rows(
        [
            {
                "congress_no": 118,
                "chamber": "H",
                "bioguide_id": "A000001",
                "committee_code": "HSAG",
                "committee_name": "Committee on Agriculture",
                "start_date": date(2023, 1, 3),
                "congress_end_date": date(2025, 1, 3),
                "internal_start_event_id": "db-1",
                "internal_end_event_id": "db-2",
                "end_effective_date": date(2023, 8, 1),
            },
            {
                "congress_no": 118,
                "chamber": "S",
                "bioguide_id": "B000002",
                "committee_code": "SSAF",
                "committee_name": "Committee on Agriculture, Nutrition, and Forestry",
                "start_date": date(2023, 1, 3),
                "congress_end_date": date(2025, 1, 3),
                "internal_start_event_id": "db-3",
                "internal_end_event_id": None,
                "end_effective_date": None,
            },
            {
                "congress_no": 118,
                "chamber": "S",
                "bioguide_id": "C000003",
                "committee_code": "SSAF",
                "committee_name": "Committee on Agriculture, Nutrition, and Forestry",
                "start_date": date(2023, 1, 3),
                "congress_end_date": date(2025, 1, 3),
                "membership_end_boundary": date(2024, 1, 10),
                "internal_start_event_id": "db-3",
                "internal_end_event_id": None,
                "end_effective_date": None,
            },
        ],
        release_version="v0.1.0",
        event_id_map=event_id_map,
    )

    early = rows[0]
    assert early["start_release_event_id"] == event_id_map["db-1"]
    assert early["end_release_event_id"] == event_id_map["db-2"]
    assert early["last_active_date"] == date(2023, 7, 31)
    assert early["termination_effective_date"] == date(2023, 8, 1)
    assert early["ended_early"] is True

    natural = rows[1]
    assert natural["start_release_event_id"] == event_id_map["db-3"]
    assert natural["end_release_event_id"] is None
    assert natural["last_active_date"] == date(2025, 1, 2)
    assert natural["termination_effective_date"] == date(2025, 1, 3)
    assert natural["ended_early"] is False

    closed_without_end_event = rows[2]
    assert closed_without_end_event["end_release_event_id"] is None
    assert closed_without_end_event["termination_effective_date"] == date(2024, 1, 10)
    assert closed_without_end_event["last_active_date"] == date(2024, 1, 9)
    assert closed_without_end_event["ended_early"] is True


def test_source_runtime_timestamps_are_omitted_from_canonical_exports() -> None:
    rows = build_source_rows(
        [
            {
                "release_source_id": "release-source:test",
                "internal_source_id": 5,
                "release_source_document_id": "release-source-document:test",
                "internal_source_document_id": 11,
                "source_type": "resolution",
                "source_name": "House Resolution",
                "version_tag": "2023-01-03",
                "external_id": "HRES-1",
                "doc_date": date(2023, 1, 3),
                "url": "https://example.test/HRES-1",
                "content_hash": "abc123",
                "retrieved_at_utc": datetime(2026, 7, 14, 12, 0, 0),
                "created_at_utc": datetime(2026, 7, 14, 13, 0, 0),
            }
        ],
        release_version="v0.1.0",
    )
    assert rows[0]["retrieved_at_utc"] is None
    assert rows[0]["created_at_utc"] is None


def test_stable_now_honors_source_date_epoch(monkeypatch) -> None:
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "1784246400")

    assert _stable_now() == "2026-07-17T00:00:00Z"


def test_scope_filter_excludes_joint_and_select_special() -> None:
    assert is_standing_committee(chamber="H", committee_code="HSAG", committee_name="Committee on Agriculture")
    assert not is_standing_committee(chamber="H", committee_code="HLIG", committee_name="Select Committee on Intelligence")
    assert not is_standing_committee(chamber="S", committee_code="SSPC", committee_name="Special Committee on Aging")
    assert not is_standing_committee(chamber="J", committee_code="JSEC", committee_name="Joint Economic Committee", is_joint=True)


def test_csv_serialization_is_utf8_and_stable() -> None:
    spec = DATASET_SPECS["committees"]
    rows = [
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "committee_code": "HSAG",
            "chamber": "H",
            "committee_name": "Comité on Agriculture",
            "valid_start": date(2011, 1, 5),
            "valid_last_active_date": None,
            "is_joint": False,
        }
    ]
    serialized = serialize_rows_for_csv(spec, rows)
    assert list(serialized[0]) == [column.name for column in spec.columns]
    encoded = ",".join(serialized[0].values()).encode("utf-8")
    assert encoded.decode("utf-8").endswith("false")
    assert "Comité" in encoded.decode("utf-8")


def test_workbook_parity_and_typed_dates(tmp_path: Path) -> None:
    sheet_rows = _sample_sheet_rows()
    semantic_hash = compute_semantic_workbook_hash(sheet_rows)
    sheet_rows["Release Metadata"] = build_release_metadata_rows(
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "generated_at_utc": "2026-07-14T00:00:00Z",
            "workbook_semantic_sha256": semantic_hash,
            "scope": "standing",
        }
    )
    path = tmp_path / "release.xlsx"
    write_workbook(path, sheet_rows)

    reopened = load_workbook(path, read_only=False, data_only=False)
    assert reopened.sheetnames == [
        "Assignments",
        "Rankings",
        "Events",
        "Members",
        "Committees",
        "Sources",
        "Validation",
        "Data Dictionary",
        "Release Metadata",
    ]
    assignments_sheet = reopened["Assignments"]
    assert assignments_sheet.freeze_panes == "A2"
    assert assignments_sheet["H2"].value == datetime(2023, 1, 3)
    assert assignments_sheet["I2"].value == datetime(2023, 7, 31)
    assert assignments_sheet["J2"].value == datetime(2023, 8, 1)
    assert assignments_sheet.auto_filter.ref == assignments_sheet.dimensions

    read_only = load_workbook(path, read_only=True, data_only=False)
    first_row = next(read_only["Assignments"].iter_rows(min_row=1, max_row=1, values_only=True))
    assert first_row[0] == "release_version"

    extracted = extract_sheet_rows_from_workbook(path)
    for expected_sheet in WORKBOOK_SHEET_ORDER:
        spec = next(
            value for value in DATASET_SPECS.values() if value.sheet_name == expected_sheet
        )
        original = serialize_rows_for_csv(spec, sheet_rows[expected_sheet])
        round_tripped = serialize_rows_for_csv(spec, extracted[expected_sheet])
        assert round_tripped == original


def test_release_artifact_writer_includes_rankings_sheet(tmp_path: Path) -> None:
    sheet_rows = _sample_sheet_rows()
    file_rows = {
        key: sheet_rows[spec.sheet_name]
        for key, spec in DATASET_SPECS.items()
        if spec.sheet_name in sheet_rows
    }
    file_rows["directory_mismatches"] = []

    write_release_artifacts(
        tmp_path,
        file_rows,
        congress_from=118,
        congress_to=118,
        release_metadata={"schema_version": SCHEMA_VERSION},
    )

    workbook = load_workbook(
        tmp_path / "committee_membership_118_118.xlsx",
        read_only=True,
        data_only=False,
    )
    assert workbook.sheetnames == list(WORKBOOK_SHEET_ORDER)
    assert workbook["Rankings"].max_row == 2


def test_workbook_bytes_are_deterministic(tmp_path: Path) -> None:
    sheet_rows = _sample_sheet_rows()
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"

    write_workbook(first, sheet_rows)
    write_workbook(second, sheet_rows)

    assert first.read_bytes() == second.read_bytes()
    with zipfile.ZipFile(first) as workbook_archive:
        core_properties = workbook_archive.read("docProps/core.xml")
    assert (
        b'<dcterms:modified xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>'
        in core_properties
    )
    assert b"<dc:creator>Committee Steward</dc:creator>" in core_properties
    assert b"<cp:lastModifiedBy>Committee Steward</cp:lastModifiedBy>" in core_properties
    assert b"Provider" not in core_properties


def test_semantic_hash_ignores_volatile_metadata_field_values() -> None:
    base = _sample_sheet_rows()
    changed = _sample_sheet_rows()
    base["Release Metadata"] = build_release_metadata_rows(
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "generated_at_utc": "2026-07-14T00:00:00Z",
            "workbook_semantic_sha256": "",
            "scope": "standing",
        }
    )
    changed["Release Metadata"] = build_release_metadata_rows(
        {
            "release_version": "v0.1.0",
            "schema_version": SCHEMA_VERSION,
            "generated_at_utc": "2026-08-01T12:34:56Z",
            "workbook_semantic_sha256": "different",
            "scope": "standing",
        }
    )

    assert compute_semantic_workbook_hash(base) == compute_semantic_workbook_hash(changed)


def test_directory_mismatch_release_rows_keep_typed_snapshot_dates() -> None:
    snapshot_date = date(2024, 4, 25)
    rows = build_directory_mismatch_rows(
        [
            DirectoryMismatch(
                congress_no=118,
                snapshot_date=snapshot_date,
                chamber="H",
                committee_scope="standing",
                side="directory_only",
                raw_member_name="Example Member",
                bioguide_id="E000001",
                committee_text="Agriculture",
                committee_code="HSAG",
                detail="",
            )
        ],
        release_version="v0.1.0",
    )

    assert rows[0]["snapshot_date"] == snapshot_date
